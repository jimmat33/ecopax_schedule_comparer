from openpyxl import *
import string
import tkinter as tk
from tkinter import *
import tkinter.font
import os
from tkinter.tix import Select
from tkinter.ttk import *
from  tkinter import ttk
from tkcalendar import Calendar, DateEntry
from tkinter import filedialog
import xlsxwriter


class ScheduleGUI(object):
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Schedule Change Report Generator")
        self.root.geometry('525x200')
        self.root.resizable(False, False)
        img = tk.PhotoImage(file= (os.path.abspath('gui_icon.png')))
        self.root.tk.call('wm', 'iconphoto', self.root._w, img)
        self.original_sheet_fp = ''
        self.new_sheet_fp = ''
        self.orig_sheet_label = 0
        self.new_sheet_label = 0
        self.change_lst = []

    def run_gui(self):
        self.init_widgets()
        self.root.mainloop()

    def init_widgets(self):
        large_widget_font = tk.font.Font(family = 'MS Shell Dlg 2', size = 14)

        self.import_orig_sheet_label = Label(self.root, text= 'Import Original Sheet:', font = large_widget_font, state = 'normal')
        self.import_orig_sheet_label.pack()

        self.import_orig_sheet_button = Button(self.root, text = 'Import Spreadsheet', state = 'normal', command = self.import_orig_spreadsheet_btn_click)
        self.import_orig_sheet_button.pack()
        self.orig_sheet_label = Label(self.root, text= '', state = 'normal')
        self.orig_sheet_label.pack()

        self.import_new_sheet_label = Label(self.root, text= 'Import New Sheet:', font = large_widget_font, state = 'normal')
        self.import_new_sheet_label.pack()

        self.import_new_sheet_button = Button(self.root, text = 'Import Spreadsheet', state = 'normal', command = self.import_new_spreadsheet_btn_click)
        self.import_new_sheet_button.pack()
        self.new_sheet_label = Label(self.root, text= '', state = 'normal')
        self.new_sheet_label.pack()

        self.find_changes_label = Label(self.root, text= 'Get Changes:', font = large_widget_font, state = 'normal')
        self.find_changes_label.pack()
        self.find_changes_button = Button(self.root, text = 'Generate Change Report', state = 'normal', command = self.compare_sheets)
        self.find_changes_button.pack()


    def import_orig_spreadsheet_btn_click(self):
        filename = filedialog.askopenfilenames(initialdir = "", title = "Select a File", filetypes = (("all files", "*.*"),))

        self.original_sheet_fp = filename[0]
        self.orig_sheet_label['text'] = filename[0]

        self.change_lst.clear()

    def import_new_spreadsheet_btn_click(self):
        new_filename = filedialog.askopenfilenames(initialdir = "", title = "Select a File", filetypes = (("all files", "*.*"),))

        self.new_sheet_fp = new_filename[0]
        self.new_sheet_label['text'] = new_filename[0]
        self.change_lst.clear()

    def create_change_sheet(self):
        desktop = str(os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')) + '\\schedule-changelog.xlsx'
        workbook = xlsxwriter.Workbook(desktop)
        worksheet = workbook.add_worksheet()
        
        i = 0
        while i < len(self.change_lst):
            cell_index = f'A{i}'
            worksheet.write(cell_index, self.change_lst[i])

            i += 1

        workbook.close()

    def compare_sheets(self):
        if self.original_sheet_fp != '' and self.new_sheet_fp != '':
        
            wb1 = load_workbook(filename=self.original_sheet_fp)
            wb2 = load_workbook(filename=self.new_sheet_fp)

            sh1 = wb1.active
            sh2 = wb2.active 

            max_row = sh2.max_row
            max_column = sh2.max_column

            i = 1
            j = 1

            while i < max_row:
                j = 1
                while j < max_column:
                    if sh1.cell(row = i, column = j).value != sh2.cell(row = i, column = j).value:
                        self.change_lst.append(f'{sh1.cell(row = i, column = 1).value} for the {sh1.cell(row = 2, column = j).value} shift changed from {sh1.cell(row = i, column = j).value} --> {sh2.cell(row = i, column = j).value}')
                    j += 1
                i += 1

        self.create_change_sheet()
